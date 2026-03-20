"""
Shopee Affiliate Marketing Automation — Orchestrator

Usage:
  python main.py                  # Full pipeline: scrape + AI copy + save Excel
  python main.py --scheduler      # Start post scheduler daemon only
  python main.py --dashboard      # Launch Streamlit dashboard only
  python main.py --all            # Full pipeline + scheduler daemon
  python main.py --dry-run        # Validate config + imports, no real calls
"""

import argparse
import asyncio
import logging
import os
import subprocess
import sys
import threading
from datetime import datetime
from pathlib import Path

import pandas as pd

# ── Bootstrap logging before any imports that use it ────────────────────────
def _setup_logging(log_path: str):
    Path(log_path).parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )

_setup_logging("logs/affiliate_bot.log")
logger = logging.getLogger(__name__)


# ── Validação de licença ─────────────────────────────────────────────────────
def _check_license():
    """Valida licença antes de iniciar o bot. Exibe tela de ativação se necessário."""
    try:
        sys.path.insert(0, str(Path(__file__).parent / "bot_client"))
        from license_validator import get_saved_license, validate_license

        license_key = get_saved_license()

        if license_key:
            valid, msg = validate_license(license_key)
            if valid:
                logger.info(f"Licença OK: {msg}")
                return
            logger.warning(f"Licença inválida: {msg}")

        # Exibe tela de ativação
        from activation_ui import show_activation_window
        key = show_activation_window()
        if not key:
            logger.error("Ativação cancelada pelo usuário.")
            sys.exit(1)

        valid, msg = validate_license(key)
        if not valid:
            logger.error(f"Ativação falhou: {msg}")
            sys.exit(1)

        logger.info(f"Bot ativado com sucesso: {msg}")

    except ImportError:
        # Em desenvolvimento local (sem bot_client), pula a validação
        logger.warning("Módulo de licença não encontrado — modo desenvolvimento.")


_check_license()

# Excel concurrency lock — shared with scheduler
excel_lock = threading.Lock()


def save_to_excel(products: list, filepath: str) -> str:
    """Save product list to Excel with auto-formatted columns."""
    from config.settings import settings
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)

    # Add default scheduling fields + ganho estimado
    for p in products:
        p.setdefault("status_agendamento", "pendente")
        p.setdefault("data_publicacao", "")
        p.setdefault("preco", 0)
        # Ganho estimado = preço × (comissão% / 100)
        preco = float(p.get("preco") or 0)
        comissao = float(p.get("comissao_novo") or 0)
        p["ganho_estimado"] = round(preco * comissao / 100, 2) if preco > 0 else 0

    df = pd.DataFrame(products)

    # Keep only expected columns (add missing ones as empty)
    for col in settings.EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[list(settings.EXCEL_COLUMNS)]

    with excel_lock:
        df.to_excel(filepath, index=False, engine="openpyxl")

        # Auto-fit columns and style header
        wb = load_workbook(filepath)
        ws = wb.active
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(settings.EXCEL_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}1"].fill = header_fill
            ws[f"{col_letter}1"].font = header_font
            # Auto-fit width
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Highlight AI-failed rows in orange
        orange_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
        ai_col = list(settings.EXCEL_COLUMNS).index("overlay") + 1
        for row in range(2, ws.max_row + 1):
            status_col = list(settings.EXCEL_COLUMNS).index("status_agendamento") + 1
            # Check if ai_status is failed via overlay being the fallback
            overlay_val = str(ws.cell(row=row, column=ai_col).value or "")
            if overlay_val == "Achou incrível aqui 🔥":
                for c in range(1, len(settings.EXCEL_COLUMNS) + 1):
                    ws.cell(row=row, column=c).fill = orange_fill

        wb.save(filepath)

    abs_path = str(Path(filepath).resolve())
    logger.info(f"Excel salvo: {abs_path} ({len(products)} produtos)")
    return abs_path


async def run_pipeline(settings, dry_run: bool = False) -> list:
    """Main pipeline: scrape → AI copy → save Excel."""
    from scraper.shopee_affiliate import scrape_all_products
    from ai.copy_generator import generate_batch

    logger.info("=" * 60)
    logger.info("INICIANDO PIPELINE")
    logger.info(f"Modo: {'DRY-RUN' if dry_run else 'PRODUÇÃO'}")
    logger.info(f"Máx produtos: {settings.max_products}")
    logger.info("=" * 60)

    if dry_run:
        logger.info("[DRY-RUN] Gerando dados fictícios para teste...")
        products = _mock_products(5)
    else:
        logger.info("Etapa 1/3 — Scraping Shopee Affiliate...")
        products = await scrape_all_products(settings)
        if not products:
            logger.error("Nenhum produto coletado. Verifique login e seletores.")
            return []

    logger.info(f"Etapa 2/3 — Gerando copy com Claude ({len(products)} produtos)...")
    if dry_run:
        products = _mock_copy(products)
    else:
        products = generate_batch(
            products,
            api_key=settings.anthropic_api_key,
            model=settings.anthropic_model,
        )

    logger.info("Etapa 3/3 — Salvando Excel...")
    save_to_excel(products, settings.EXCEL_PATH)

    logger.info("Pipeline concluído!")
    return products


def _mock_products(n: int) -> list:
    """Generate realistic mock products covering different categories for dry-run."""
    _samples = [
        ("Fone Bluetooth JBL TWS sem fio preto",           "eletronicos", 89.90,  14.0),
        ("Serum Vitamina C facial anti-idade 30ml",        "beleza",      45.00,  18.0),
        ("Vestido floral manga longa feminino P-GG",       "moda",        79.90,  12.0),
        ("Airfryer Digital 4L 1400W inox",                 "casa",       189.90,  10.0),
        ("Whey Protein Chocolate 1kg Growth",              "saude",       89.90,  11.0),
        ("Coleira antipulga gato cachorro ajustavel",      "pet",         29.90,  20.0),
        ("Kit berco bebe menino 9 pecas algodao",          "bebe",       149.90,   9.0),
        ("Suporte articulado celular carro 360 graus",     "eletronicos", 24.90,  16.0),
    ]
    produtos = []
    for i in range(min(n, len(_samples))):
        nome, cat, preco, comm = _samples[i]
        produtos.append({
            "produto":       nome,
            "product_id":    f"1000{i}",
            "link_original": f"https://shopee.com.br/{cat}-produto-i.{12345+i}.{i+1}",
            "link_afiliado": f"https://s.shopee.com.br/dryrun{i+1}",
            "comissao_novo": comm,
            "comissao_atual": round(comm * 0.6, 1),
            "preco":          preco,
            "thumbnail_url":  "",
            "category":       cat,
        })
    return produtos


def _mock_copy(products: list) -> list:
    """
    Dry-run: chama o gerador de copy real (com categorias) mas sem API Claude.
    Usa os fallbacks por categoria para mostrar copy realista no dashboard.
    """
    from ai.copy_generator import detect_category, FALLBACK_COPIES, CATEGORY_LABELS

    for p in products:
        cat  = detect_category(p.get("produto", ""), p.get("link_original", ""))
        link = p.get("link_afiliado", "")
        fb   = FALLBACK_COPIES.get(cat, FALLBACK_COPIES["geral"])
        nome = p.get("produto", "produto")
        preco = p.get("preco", 0)
        preco_txt = f"R$ {preco:.2f}" if preco > 0 else "preco imperdivel"

        p.update({
            "overlay":             fb["overlay"],
            "legenda":             fb["legenda"].format(link=link),
            "hashtags":            fb["hashtags"],
            "estrategia":          f"[DRY-RUN] {fb['estrategia']}",
            "ai_status":           "dry_run",
            "categoria_detectada": cat,
        })
    return products


def start_scheduler(settings):
    """Start the APScheduler post daemon."""
    from scheduler.post_scheduler import create_and_start_scheduler
    scheduler = create_and_start_scheduler(settings, excel_lock)
    logger.info("Scheduler iniciado. Pressione Ctrl+C para parar.")
    try:
        while True:
            import time
            time.sleep(60)
    except KeyboardInterrupt:
        scheduler.shutdown()
        logger.info("Scheduler encerrado.")


def start_dashboard():
    """Launch the Streamlit dashboard in a subprocess."""
    logger.info("Iniciando dashboard em http://localhost:8501")
    subprocess.run([sys.executable, "-m", "streamlit", "run", "dashboard/app.py"])


def main():
    parser = argparse.ArgumentParser(description="Shopee Affiliate Automation Bot")
    parser.add_argument("--scheduler", action="store_true", help="Iniciar scheduler de postagens")
    parser.add_argument("--dashboard", action="store_true", help="Abrir dashboard Streamlit")
    parser.add_argument("--all", action="store_true", help="Pipeline + scheduler")
    parser.add_argument("--dry-run", action="store_true", help="Teste sem chamadas reais")
    args = parser.parse_args()

    # Load settings
    try:
        from config.settings import get_settings
        settings = get_settings()
    except EnvironmentError as e:
        logger.error(f"Configuração inválida: {e}")
        logger.error("Copie .env.example para .env e preencha as variáveis.")
        sys.exit(1)

    if args.dashboard:
        start_dashboard()
        return

    if args.scheduler:
        start_scheduler(settings)
        return

    # Run pipeline
    products = asyncio.run(run_pipeline(settings, dry_run=args.dry_run))

    if args.all and not args.dry_run:
        # Start scheduler in background thread after pipeline
        t = threading.Thread(target=start_scheduler, args=(settings,), daemon=True)
        t.start()
        logger.info("Pipeline concluído. Scheduler rodando em background.")
        logger.info("Pressione Ctrl+C para encerrar.")
        try:
            t.join()
        except KeyboardInterrupt:
            logger.info("Encerrando...")

    if products:
        logger.info(f"\nResumo: {len(products)} produtos processados.")
        logger.info(f"Excel: {Path(settings.EXCEL_PATH).resolve()}")
        logger.info("Para ver o dashboard: python main.py --dashboard")


if __name__ == "__main__":
    main()
